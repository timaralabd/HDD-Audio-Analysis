# ============================================================
# Görev 6.1 - Ses Verisi Çoğaltma (Data Augmentation)
# Pitch Shifting, Time Stretching ve Background Noise
# ============================================================

import os
import re
import numpy as np
import pandas as pd
import librosa
import soundfile as sf

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


print("\nGörev 6.1 Ses Verisi Çoğaltma başladı...")


# ------------------------------------------------------------
# Klasörler
# ------------------------------------------------------------

INPUT_DIR = "gorev_4_3_standardized"
OUTPUT_DIR = "gorev_6_1_augmented"

os.makedirs(OUTPUT_DIR, exist_ok=True)

np.random.seed(42)


# ------------------------------------------------------------
# Etiket dosyasını oku
# ------------------------------------------------------------

feature_files = [
    "gorev_5_1_feature_matrix_duzenli.csv",
    "gorev_5_1_feature_matrix.csv"
]

feature_file = None

for candidate in feature_files:
    if os.path.exists(candidate):
        feature_file = candidate
        break

if feature_file is None:
    raise FileNotFoundError(
        "Feature Matrix CSV dosyası bulunamadı."
    )

feature_df = pd.read_csv(feature_file)


# Sütun isimlerini otomatik bul
if "File" in feature_df.columns:
    file_column = "File"
elif "dosya" in feature_df.columns:
    file_column = "dosya"
else:
    raise ValueError("Dosya sütunu bulunamadı.")

if "Class" in feature_df.columns:
    class_column = "Class"
elif "hata_sinifi" in feature_df.columns:
    class_column = "hata_sinifi"
else:
    raise ValueError("Hata sınıfı sütunu bulunamadı.")


# Dosya numarası -> sınıf eşleştirmesi
label_dictionary = {}

for _, row in feature_df.iterrows():

    filename = str(row[file_column])
    match = re.search(r"(\d+)(?=\.wav|\.mp3|$)", filename)

    if match:
        file_number = int(match.group(1))
        label_dictionary[file_number] = str(row[class_column])


# ------------------------------------------------------------
# Yardımcı fonksiyonlar
# ------------------------------------------------------------

def dosya_numarasini_bul(filename):
    """Dosya adından Audio Player numarasını bulur."""

    match = re.search(r"(\d+)(?=\.wav$)", filename)

    if match:
        return int(match.group(1))

    return None


def sabit_uzunluk(y, target_length):
    """Sesi orijinal örnek sayısına getirir."""

    if len(y) > target_length:
        return y[:target_length]

    if len(y) < target_length:
        return np.pad(
            y,
            (0, target_length - len(y)),
            mode="constant"
        )

    return y


def sesi_kaydet(path, y, sample_rate):
    """Sesi taşma olmadan WAV olarak kaydeder."""

    y = np.nan_to_num(y)
    y = np.clip(y, -1.0, 1.0)

    sf.write(
        path,
        y,
        sample_rate,
        subtype="PCM_16"
    )


def arka_plan_gurultusu_ekle(y, noise_level=0.008):
    """Sese düşük seviyeli arka plan gürültüsü ekler."""

    noise = np.random.normal(
        0,
        1,
        len(y)
    )

    noise = noise * noise_level

    return y + noise


# ------------------------------------------------------------
# Augmentation işlemi
# ------------------------------------------------------------

report_rows = []

wav_files = sorted([
    file for file in os.listdir(INPUT_DIR)
    if file.lower().endswith(".wav")
])

print("Orijinal ses sayısı:", len(wav_files))


for file in wav_files:

    input_path = os.path.join(INPUT_DIR, file)

    y, sample_rate = librosa.load(
        input_path,
        sr=None,
        mono=True
    )

    original_length = len(y)
    duration = original_length / sample_rate

    file_number = dosya_numarasini_bul(file)
    class_name = label_dictionary.get(
        file_number,
        "Bilinmiyor"
    )

    base_name = os.path.splitext(file)[0]

    # --------------------------------------------------------
    # 1. Pitch Shift: +2 semitone
    # --------------------------------------------------------

    pitch_audio = librosa.effects.pitch_shift(
        y=y,
        sr=sample_rate,
        n_steps=2
    )

    pitch_audio = sabit_uzunluk(
        pitch_audio,
        original_length
    )

    pitch_name = f"{base_name}_pitch_plus2.wav"
    pitch_path = os.path.join(OUTPUT_DIR, pitch_name)

    sesi_kaydet(
        pitch_path,
        pitch_audio,
        sample_rate
    )

    report_rows.append({
        "Kaynak Dosya": file,
        "Yeni Dosya": pitch_name,
        "Hata Sınıfı": class_name,
        "Augmentation Türü": "Pitch Shift (+2)",
        "Süre (sn)": round(duration, 3)
    })

    # --------------------------------------------------------
    # 2. Time Stretch: %10 yavaşlatma
    # --------------------------------------------------------

    stretch_audio = librosa.effects.time_stretch(
        y=y,
        rate=0.90
    )

    stretch_audio = sabit_uzunluk(
        stretch_audio,
        original_length
    )

    stretch_name = f"{base_name}_time_stretch_090.wav"
    stretch_path = os.path.join(
        OUTPUT_DIR,
        stretch_name
    )

    sesi_kaydet(
        stretch_path,
        stretch_audio,
        sample_rate
    )

    report_rows.append({
        "Kaynak Dosya": file,
        "Yeni Dosya": stretch_name,
        "Hata Sınıfı": class_name,
        "Augmentation Türü": "Time Stretch (0.90)",
        "Süre (sn)": round(duration, 3)
    })

    # --------------------------------------------------------
    # 3. Background Noise
    # --------------------------------------------------------

    noise_audio = arka_plan_gurultusu_ekle(
        y,
        noise_level=0.008
    )

    noise_audio = sabit_uzunluk(
        noise_audio,
        original_length
    )

    noise_name = f"{base_name}_background_noise.wav"
    noise_path = os.path.join(
        OUTPUT_DIR,
        noise_name
    )

    sesi_kaydet(
        noise_path,
        noise_audio,
        sample_rate
    )

    report_rows.append({
        "Kaynak Dosya": file,
        "Yeni Dosya": noise_name,
        "Hata Sınıfı": class_name,
        "Augmentation Türü": "Background Noise",
        "Süre (sn)": round(duration, 3)
    })

    print("Çoğaltıldı:", file)


# ------------------------------------------------------------
# Rapor oluştur
# ------------------------------------------------------------

report_df = pd.DataFrame(report_rows)

report_df.to_csv(
    "gorev_6_1_augmentation_raporu.csv",
    index=False
)

excel_file = "gorev_6_1_augmentation_raporu.xlsx"

with pd.ExcelWriter(
    excel_file,
    engine="openpyxl"
) as writer:

    report_df.to_excel(
        writer,
        sheet_name="Augmentation Report",
        index=False
    )

    ws = writer.sheets["Augmentation Report"]
    ws.freeze_panes = "A2"

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center"
        )

    for column_cells in ws.columns:

        max_length = 0

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        column_letter = get_column_letter(
            column_cells[0].column
        )

        ws.column_dimensions[
            column_letter
        ].width = max_length + 3


print("\nGörev 6.1 tamamlandı.")
print("Orijinal dosya sayısı:", len(wav_files))
print("Üretilen yeni dosya sayısı:", len(report_df))
print("Yeni ses klasörü:", OUTPUT_DIR)
print("Excel raporu:", excel_file)