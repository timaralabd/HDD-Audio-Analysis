# ============================================================
# Görev 6.1 - Ses Verisi Çoğaltma (Data Augmentation)
# Librosa kullanmaz: Python 3.14 ile daha uyumlu sürüm
# Pitch Shifting, Time Stretching ve Background Noise
# ============================================================

import os
import re
import numpy as np
import pandas as pd

from scipy.io import wavfile
from scipy.signal import resample
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("\nGörev 6.1 Ses Verisi Çoğaltma başladı...")

INPUT_DIR = "gorev_4_3_standardized"
OUTPUT_DIR = "gorev_6_1_augmented"
os.makedirs(OUTPUT_DIR, exist_ok=True)

np.random.seed(42)

feature_candidates = [
    "gorev_5_1_feature_matrix_duzenli.csv",
    "gorev_5_1_feature_matrix.csv"
]

feature_file = next(
    (name for name in feature_candidates if os.path.exists(name)),
    None
)

if feature_file is None:
    raise FileNotFoundError(
        "Feature Matrix CSV bulunamadı. "
        "gorev_5_1_feature_matrix_duzenli.csv veya "
        "gorev_5_1_feature_matrix.csv dosyasını kontrol et."
    )

feature_df = pd.read_csv(feature_file)

if "File" in feature_df.columns:
    file_column = "File"
elif "dosya" in feature_df.columns:
    file_column = "dosya"
else:
    raise ValueError("Dosya sütunu bulunamadı.")

if "Class" in feature_df.columns:
    class_column = "Class"
elif "hata_sinifi" in feature_df.columns:
    class_column = "hata_sinifi"
else:
    raise ValueError("Hata sınıfı sütunu bulunamadı.")

label_dictionary = {}

for _, row in feature_df.iterrows():
    filename = str(row[file_column])
    match = re.search(r"(\d+)(?=\.(?:wav|mp3)$)", filename)

    if match:
        label_dictionary[int(match.group(1))] = str(row[class_column])


def dosya_numarasini_bul(filename):
    match = re.search(r"(\d+)(?=\.wav$)", filename)
    return int(match.group(1)) if match else None


def mono_yap(data):
    if data.ndim > 1:
        data = data[:, 0]
    return data.astype(np.float64)


def normalize_int16(data):
    data = np.nan_to_num(data)

    max_value = np.max(np.abs(data))
    if max_value > 0:
        data = data / max_value * 32767

    return np.clip(data, -32768, 32767).astype(np.int16)


def sabit_uzunluk(data, target_length):
    if len(data) > target_length:
        return data[:target_length]

    if len(data) < target_length:
        return np.pad(
            data,
            (0, target_length - len(data)),
            mode="constant"
        )

    return data


def pitch_shift(data, semitones=2):
    """
    Basit pitch shifting:
    Önce örnek sayısını perde oranına göre değiştirir,
    sonra orijinal uzunluğa getirir.
    """
    factor = 2 ** (semitones / 12)
    shifted_length = max(1, int(len(data) / factor))
    shifted = resample(data, shifted_length)
    return sabit_uzunluk(shifted, len(data))


def time_stretch(data, rate=0.90):
    """
    rate < 1: yavaşlatır
    rate > 1: hızlandırır
    """
    stretched_length = max(1, int(len(data) / rate))
    stretched = resample(data, stretched_length)
    return sabit_uzunluk(stretched, len(data))


def background_noise(data, noise_level=0.008):
    signal_std = np.std(data)

    if signal_std == 0:
        signal_std = 1.0

    noise = np.random.normal(
        loc=0.0,
        scale=signal_std * noise_level,
        size=len(data)
    )

    return data + noise


wav_files = sorted(
    file for file in os.listdir(INPUT_DIR)
    if file.lower().endswith(".wav")
)

if not wav_files:
    raise FileNotFoundError(
        f"{INPUT_DIR} klasöründe WAV dosyası bulunamadı."
    )

report_rows = []

print("Orijinal ses sayısı:", len(wav_files))

for file in wav_files:
    input_path = os.path.join(INPUT_DIR, file)

    sample_rate, data = wavfile.read(input_path)
    data = mono_yap(data)

    original_length = len(data)
    duration = original_length / sample_rate

    file_number = dosya_numarasini_bul(file)
    class_name = label_dictionary.get(file_number, "Bilinmiyor")
    base_name = os.path.splitext(file)[0]

    augmentations = [
        (
            "Pitch Shift (+2)",
            f"{base_name}_pitch_plus2.wav",
            pitch_shift(data, semitones=2)
        ),
        (
            "Time Stretch (0.90)",
            f"{base_name}_time_stretch_090.wav",
            time_stretch(data, rate=0.90)
        ),
        (
            "Background Noise",
            f"{base_name}_background_noise.wav",
            background_noise(data, noise_level=0.008)
        )
    ]

    for augmentation_type, new_name, augmented_data in augmentations:
        output_path = os.path.join(OUTPUT_DIR, new_name)

        augmented_data = sabit_uzunluk(
            augmented_data,
            original_length
        )

        wavfile.write(
            output_path,
            sample_rate,
            normalize_int16(augmented_data)
        )

        report_rows.append({
            "Kaynak Dosya": file,
            "Yeni Dosya": new_name,
            "Hata Sınıfı": class_name,
            "Augmentation Türü": augmentation_type,
            "Sample Rate": sample_rate,
            "Süre (sn)": round(duration, 3)
        })

    print("Çoğaltıldı:", file)

report_df = pd.DataFrame(report_rows)

report_df.to_csv(
    "gorev_6_1_augmentation_raporu.csv",
    index=False
)

excel_file = "gorev_6_1_augmentation_raporu.xlsx"

with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
    report_df.to_excel(
        writer,
        sheet_name="Augmentation Report",
        index=False
    )

    ws = writer.sheets["Augmentation Report"]
    ws.freeze_panes = "A2"

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for column_cells in ws.columns:
        max_length = max(
            len(str(cell.value))
            for cell in column_cells
            if cell.value is not None
        )

        column_letter = get_column_letter(
            column_cells[0].column
        )

        ws.column_dimensions[column_letter].width = max_length + 3

print("\nGörev 6.1 tamamlandı.")
print("Orijinal dosya sayısı:", len(wav_files))
print("Üretilen yeni dosya sayısı:", len(report_df))
print("Yeni ses klasörü:", OUTPUT_DIR)
print("Excel raporu:", excel_file)
