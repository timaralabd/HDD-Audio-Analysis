# ============================================================
# Görev 6.2
# Augmentation Öncesi ve Sonrası Model Karşılaştırması
# Random Forest - SVM - k-NN
# ============================================================

import os
import re
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from scipy.io import wavfile
from python_speech_features import mfcc

from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import (
    accuracy_score,
    precision_score,
    recall_score,
    f1_score
)

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


print("\nGörev 6.2 Model Karşılaştırması başladı...")


# ============================================================
# 1. Dosya ve klasör ayarları
# ============================================================

ORIGINAL_AUDIO_DIR = "gorev_4_3_standardized"
AUGMENTED_AUDIO_DIR = "gorev_6_1_augmented"

FEATURE_FILE_CANDIDATES = [
    "gorev_5_1_feature_matrix_duzenli.csv",
    "gorev_5_1_feature_matrix.csv"
]

RANDOM_STATE = 42


# ============================================================
# 2. Feature Matrix dosyasını bul
# ============================================================

feature_file = None

for candidate in FEATURE_FILE_CANDIDATES:
    if os.path.exists(candidate):
        feature_file = candidate
        break

if feature_file is None:
    raise FileNotFoundError(
        "Görev 5.1 Feature Matrix CSV dosyası bulunamadı."
    )

feature_df = pd.read_csv(feature_file)

print("Kullanılan Feature Matrix:", feature_file)


# ============================================================
# 3. Sütun isimlerini belirle
# ============================================================

if "File" in feature_df.columns:
    file_column = "File"
elif "dosya" in feature_df.columns:
    file_column = "dosya"
else:
    raise ValueError("Dosya adı sütunu bulunamadı.")

if "Class" in feature_df.columns:
    class_column = "Class"
elif "hata_sinifi" in feature_df.columns:
    class_column = "hata_sinifi"
else:
    raise ValueError("Hata sınıfı sütunu bulunamadı.")


# ============================================================
# 4. Benzer hata sınıflarını birleştir
# ============================================================

class_mapping = {
    "head/chip": "head",
    "head stick": "head",
    "motor/head": "motor",
    "motor seizure – spindle lock.": "motor",
    "motor seizure - spindle lock.": "motor"
}

feature_df[class_column] = (
    feature_df[class_column]
    .astype(str)
    .str.strip()
    .replace(class_mapping)
)


# ============================================================
# 5. Dosya numarasını bulma
# ============================================================

def dosya_numarasi_bul(filename):
    """
    Dosya isminden Audio Player numarasını çıkarır.
    Örnek:
    standardized_Audio_Player_25.wav -> 25
    """

    match = re.search(
        r"Audio_Player_(\d+)",
        str(filename),
        re.IGNORECASE
    )

    if match:
        return int(match.group(1))

    return None


# Dosya numarası -> sınıf sözlüğü
label_dictionary = {}

for _, row in feature_df.iterrows():

    file_number = dosya_numarasi_bul(
        row[file_column]
    )

    if file_number is not None:
        label_dictionary[file_number] = row[class_column]


print("\nBirleştirilmiş sınıf dağılımı:")
print(pd.Series(label_dictionary.values()).value_counts())


# ============================================================
# 6. Ses dosyasından özellik çıkarma
# ============================================================

def ses_ozelliklerini_cikar(file_path):
    """
    RMS Mean, RMS Max, ZCR ve 13 MFCC ortalamasını çıkarır.
    """

    sample_rate, data = wavfile.read(file_path)

    # Stereo ise mono yap
    if data.ndim > 1:
        data = data[:, 0]

    data = data.astype(np.float32)

    if len(data) == 0:
        raise ValueError(
            f"Ses dosyası boş: {file_path}"
        )

    # RMS
    rms_mean = np.sqrt(
        np.mean(data ** 2)
    )

    rms_max = np.max(
        np.abs(data)
    )

    # ZCR
    zero_crossings = np.where(
        np.diff(np.sign(data))
    )[0]

    zcr = len(zero_crossings) / len(data)

    # MFCC
    mfcc_values = mfcc(
        data,
        samplerate=sample_rate,
        numcep=13,
        nfft=2048
    )

    mfcc_mean = np.mean(
        mfcc_values,
        axis=0
    )

    features = [
        rms_mean,
        rms_max,
        zcr
    ]

    features.extend(
        mfcc_mean.tolist()
    )

    return features


FEATURE_NAMES = [
    "RMS Mean",
    "RMS Max",
    "ZCR"
] + [
    f"MFCC{i}"
    for i in range(1, 14)
]


# ============================================================
# 7. Orijinal seslerden Feature Matrix oluştur
# ============================================================

def orijinal_veriyi_hazirla():

    rows = []

    wav_files = sorted([
        file
        for file in os.listdir(ORIGINAL_AUDIO_DIR)
        if file.lower().endswith(".wav")
    ])

    for file in wav_files:

        file_number = dosya_numarasi_bul(file)

        if file_number not in label_dictionary:
            continue

        file_path = os.path.join(
            ORIGINAL_AUDIO_DIR,
            file
        )

        features = ses_ozelliklerini_cikar(
            file_path
        )

        row = {
            "File": file,
            "Source_ID": file_number,
            "Class": label_dictionary[file_number],
            "Data_Type": "Original"
        }

        for name, value in zip(
            FEATURE_NAMES,
            features
        ):
            row[name] = value

        rows.append(row)

    return pd.DataFrame(rows)


# ============================================================
# 8. Augmented seslerden Feature Matrix oluştur
# ============================================================

def augmented_veriyi_hazirla():

    rows = []

    wav_files = sorted([
        file
        for file in os.listdir(AUGMENTED_AUDIO_DIR)
        if file.lower().endswith(".wav")
    ])

    for file in wav_files:

        file_number = dosya_numarasi_bul(file)

        if file_number not in label_dictionary:
            continue

        file_path = os.path.join(
            AUGMENTED_AUDIO_DIR,
            file
        )

        features = ses_ozelliklerini_cikar(
            file_path
        )

        row = {
            "File": file,
            "Source_ID": file_number,
            "Class": label_dictionary[file_number],
            "Data_Type": "Augmented"
        }

        for name, value in zip(
            FEATURE_NAMES,
            features
        ):
            row[name] = value

        rows.append(row)

    return pd.DataFrame(rows)


print("\nOrijinal ses özellikleri çıkarılıyor...")

original_df = orijinal_veriyi_hazirla()

print(
    "Orijinal kayıt sayısı:",
    len(original_df)
)

print("\nAugmented ses özellikleri çıkarılıyor...")

augmented_df = augmented_veriyi_hazirla()

print(
    "Augmented kayıt sayısı:",
    len(augmented_df)
)


if original_df.empty:
    raise ValueError(
        "Orijinal seslerden veri oluşturulamadı."
    )

if augmented_df.empty:
    raise ValueError(
        "Augmented seslerden veri oluşturulamadı."
    )


# Feature Matrix dosyalarını kaydet
original_df.to_csv(
    "gorev_6_2_original_feature_matrix.csv",
    index=False
)

augmented_df.to_csv(
    "gorev_6_2_augmented_feature_matrix.csv",
    index=False
)


# ============================================================
# 9. Kaynak dosyaları Train ve Test olarak ayır
# ============================================================

# Aynı sesin augmented kopyalarının test verisine sızmaması için
# bölme işlemini Source_ID üzerinden yapıyoruz.

source_table = (
    original_df[
        ["Source_ID", "Class"]
    ]
    .drop_duplicates()
)

train_sources, test_sources = train_test_split(
    source_table,
    test_size=0.20,
    random_state=RANDOM_STATE,
    stratify=source_table["Class"]
)

train_source_ids = set(
    train_sources["Source_ID"]
)

test_source_ids = set(
    test_sources["Source_ID"]
)


# Test verisi yalnızca orijinal seslerden oluşur
test_df = original_df[
    original_df["Source_ID"].isin(
        test_source_ids
    )
].copy()


# Augmentation öncesi eğitim verisi
baseline_train_df = original_df[
    original_df["Source_ID"].isin(
        train_source_ids
    )
].copy()


# Augmentation sonrası eğitim verisi:
# Orijinal train + yalnızca train dosyalarının augmented kopyaları
augmented_train_df = pd.concat(
    [
        baseline_train_df,

        augmented_df[
            augmented_df["Source_ID"].isin(
                train_source_ids
            )
        ]
    ],
    ignore_index=True
)


print("\nVeri bölünmesi:")

print(
    "Test için ayrılan orijinal dosya:",
    len(test_df)
)

print(
    "Augmentation öncesi eğitim kaydı:",
    len(baseline_train_df)
)

print(
    "Augmentation sonrası eğitim kaydı:",
    len(augmented_train_df)
)


# ============================================================
# 10. Model tanımları
# ============================================================

def modelleri_olustur():

    return {
        "Random Forest": RandomForestClassifier(
            n_estimators=500,
            max_depth=15,
            min_samples_split=3,
            min_samples_leaf=2,
            class_weight="balanced",
            random_state=RANDOM_STATE
        ),

        "SVM": Pipeline([
            (
                "scaler",
                StandardScaler()
            ),
            (
                "model",
                SVC(
                    kernel="rbf",
                    C=10,
                    gamma="scale",
                    class_weight="balanced"
                )
            )
        ]),

        "k-NN": Pipeline([
            (
                "scaler",
                StandardScaler()
            ),
            (
                "model",
                KNeighborsClassifier(
                    n_neighbors=5,
                    weights="distance"
                )
            )
        ])
    }


# ============================================================
# 11. Modelleri eğitme ve değerlendirme
# ============================================================

def modelleri_degerlendir(
    train_df,
    test_df,
    experiment_name
):

    X_train = train_df[
        FEATURE_NAMES
    ]

    y_train = train_df["Class"]

    X_test = test_df[
        FEATURE_NAMES
    ]

    y_test = test_df["Class"]

    results = []

    predictions = []

    models = modelleri_olustur()

    for model_name, model in models.items():

        print(
            f"{experiment_name} - "
            f"{model_name} eğitiliyor..."
        )

        model.fit(
            X_train,
            y_train
        )

        y_pred = model.predict(
            X_test
        )

        accuracy = accuracy_score(
            y_test,
            y_pred
        )

        precision = precision_score(
            y_test,
            y_pred,
            average="weighted",
            zero_division=0
        )

        recall = recall_score(
            y_test,
            y_pred,
            average="weighted",
            zero_division=0
        )

        f1 = f1_score(
            y_test,
            y_pred,
            average="weighted",
            zero_division=0
        )

        results.append({
            "Deney": experiment_name,
            "Model": model_name,
            "Eğitim Kayıt Sayısı": len(train_df),
            "Test Kayıt Sayısı": len(test_df),
            "Accuracy": accuracy,
            "Precision": precision,
            "Recall": recall,
            "F1-Score": f1
        })

        for file_name, true_label, predicted_label in zip(
            test_df["File"],
            y_test,
            y_pred
        ):

            predictions.append({
                "Deney": experiment_name,
                "Model": model_name,
                "Dosya": file_name,
                "Gerçek Sınıf": true_label,
                "Tahmin Edilen Sınıf": predicted_label
            })

    return (
        pd.DataFrame(results),
        pd.DataFrame(predictions)
    )


# Augmentation öncesi
before_results, before_predictions = (
    modelleri_degerlendir(
        baseline_train_df,
        test_df,
        "Augmentation Öncesi"
    )
)


# Augmentation sonrası
after_results, after_predictions = (
    modelleri_degerlendir(
        augmented_train_df,
        test_df,
        "Augmentation Sonrası"
    )
)


comparison_df = pd.concat(
    [
        before_results,
        after_results
    ],
    ignore_index=True
)

predictions_df = pd.concat(
    [
        before_predictions,
        after_predictions
    ],
    ignore_index=True
)


# Sonuçları yuvarla
metric_columns = [
    "Accuracy",
    "Precision",
    "Recall",
    "F1-Score"
]

comparison_df[metric_columns] = (
    comparison_df[metric_columns]
    .round(4)
)


print("\nModel karşılaştırma sonuçları:")
print(comparison_df)


# ============================================================
# 12. Excel raporu oluştur
# ============================================================

excel_file = "gorev_6_2_model_karsilastirma_sonuclari.xlsx"

with pd.ExcelWriter(
    excel_file,
    engine="openpyxl"
) as writer:

    comparison_df.to_excel(
        writer,
        sheet_name="Model Comparison",
        index=False
    )

    predictions_df.to_excel(
        writer,
        sheet_name="Predictions",
        index=False
    )

    original_df.to_excel(
        writer,
        sheet_name="Original Features",
        index=False
    )

    augmented_df.to_excel(
        writer,
        sheet_name="Augmented Features",
        index=False
    )

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    for ws in writer.sheets.values():

        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center"
            )

        for column_cells in ws.columns:

            max_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            column_letter = get_column_letter(
                column_cells[0].column
            )

            ws.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                45
            )


# ============================================================
# 13. Accuracy karşılaştırma grafiği
# ============================================================

accuracy_pivot = comparison_df.pivot(
    index="Model",
    columns="Deney",
    values="Accuracy"
)

accuracy_pivot.plot(
    kind="bar",
    figsize=(10, 6)
)

plt.title(
    "Augmentation Öncesi ve Sonrası Accuracy Karşılaştırması"
)

plt.xlabel("Model")
plt.ylabel("Accuracy")
plt.ylim(0, 1)
plt.xticks(rotation=0)
plt.legend(title="Deney")
plt.tight_layout()

plt.savefig(
    "gorev_6_2_accuracy_karsilastirma.png",
    dpi=300
)

plt.close()


# ============================================================
# 14. F1-Score karşılaştırma grafiği
# ============================================================

f1_pivot = comparison_df.pivot(
    index="Model",
    columns="Deney",
    values="F1-Score"
)

f1_pivot.plot(
    kind="bar",
    figsize=(10, 6)
)

plt.title(
    "Augmentation Öncesi ve Sonrası F1-Score Karşılaştırması"
)

plt.xlabel("Model")
plt.ylabel("F1-Score")
plt.ylim(0, 1)
plt.xticks(rotation=0)
plt.legend(title="Deney")
plt.tight_layout()

plt.savefig(
    "gorev_6_2_f1_karsilastirma.png",
    dpi=300
)

plt.close()


# ============================================================
# 15. Sonuç
# ============================================================

print("\nGörev 6.2 tamamlandı.")

print(
    "Excel raporu:",
    excel_file
)

print(
    "Accuracy grafiği:",
    "gorev_6_2_accuracy_karsilastirma.png"
)

print(
    "F1 grafiği:",
    "gorev_6_2_f1_karsilastirma.png"
)